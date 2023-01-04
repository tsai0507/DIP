import numpy as np
import cv2
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from PIL import Image

kid="kid blurred-noisy.tif"
fruit="fruit blurred-noisy.tif"

def allworkinIMG(img,ord_excel,weight):
    img_origin = cv2.imread(img)

    ##laplacian
    Blur_img = cv2.GaussianBlur(img_origin, (11, 11), 0)
    dst = cv2.Laplacian(Blur_img, cv2.CV_64F, ksize=3)
    img_laplacian= cv2.convertScaleAbs(dst)
    #Laplacian-sharpened
    img_laplacian_sharped=cv2.addWeighted(img_origin,1,img_laplacian,1.3,0)

    ##Sobel-gradient
    sobelx = cv2.Sobel(Blur_img,cv2.CV_64F,1,0,ksize=3)
    sobely = cv2.Sobel(Blur_img,cv2.CV_64F,0,1,ksize=3)
    img_sobel=cv2.addWeighted(sobelx,0.5,sobely,0.5,0)
    img_sobel= cv2.convertScaleAbs(img_sobel)
    img_sobel_smooth=cv2.boxFilter(img_sobel,-1, (5, 5), normalize=True)

    ##laplacian*sobel
    lap_sob=img_laplacian*img_sobel_smooth
    lap_sob= cv2.convertScaleAbs(lap_sob)

    ##kid_final
    img_final=cv2.addWeighted(img_origin,weight,lap_sob,1-weight,0)
    img_final_powerlow = np.array(255*(img_final/255)**0.6,dtype='uint8') #gamma_point_four 


    def contrast_stretching(img):
        norm_img = cv2.normalize(img, None, alpha=0,
                            beta=1, norm_type=cv2.NORM_MINMAX, dtype=cv2.CV_64F)
        norm_img = (255*norm_img).astype(np.uint8)
        return  norm_img 
    cv2.imshow('img_origin',contrast_stretching(img_origin))
    cv2.imshow('img_laplacian', contrast_stretching(img_laplacian))
    cv2.imshow('img_laplacian_sharped', contrast_stretching(img_laplacian_sharped))
    cv2.imshow('img_sobel ', contrast_stretching(img_sobel))
    cv2.imshow('img_sobel_smooth ', contrast_stretching(img_sobel_smooth))
    cv2.imshow('lap_sob ',contrast_stretching(lap_sob))
    cv2.imshow('img_final ',contrast_stretching(img_final))
    cv2.imshow('img_final_powerlow ',contrast_stretching(img_final_powerlow))
    cv2.waitKey(0)
    cv2.destroyAllWindows()

    hist = cv2.calcHist([contrast_stretching(img_origin)], [0], None, [256], [0, 255]) #convert images to histogram
    hist2 = cv2.calcHist([contrast_stretching(img_final_powerlow)], [0], None, [256], [0, 256])
    plt.title("Histogram")
    plt.xlabel("gray levels")
    plt.ylabel("pixel numbers")
    plt.plot(hist)
    plt.show()
    plt.plot(hist2)
    plt.show()

    wb = load_workbook("Histograms.xlsx")
    sheet = wb.worksheets[0]

    if(ord_excel==1):
        for i in range(2, sheet.max_row+1):
            sheet.cell(row = i, column = 2).value = hist[i-2][0]

        for i in range(2, sheet.max_row+1):
            sheet.cell(row = i, column = 3).value = hist2[i-2][0]
    elif(ord_excel==2):
        for i in range(2, sheet.max_row+1):
            sheet.cell(row = i, column = 4).value = hist[i-2][0]

        for i in range(2, sheet.max_row+1):
            sheet.cell(row = i, column = 5).value = hist2[i-2][0]

    wb.save("Histograms.xlsx")

    #origin
    output_img_origin=Image.fromarray(contrast_stretching(img_origin))
    output_img_origin.save("img/{}2original.png".format(img[:-4]),dpi=(200,200))
    #laplacian
    output_img_laplacian=Image.fromarray(contrast_stretching(img_laplacian))
    output_img_laplacian.save("img/{}2Laplacian.png".format(img[:-4]),dpi=(200,200))
    #Laplacian-sharpened
    output_img_laplacian_sharped=Image.fromarray(contrast_stretching(img_laplacian_sharped))
    output_img_laplacian_sharped.save("img/{}2Laplacian_sharped.png".format(img[:-4]),dpi=(200,200))
    #Sobel-gradient
    output_img_sobel=Image.fromarray(contrast_stretching(img_sobel))
    output_img_sobel.save("img/{}2sobel.png".format(img[:-4]),dpi=(200,200))
    #Sobel-gradient smooth
    output_img_sobel_smooth=Image.fromarray(contrast_stretching(img_sobel_smooth))
    output_img_sobel_smooth.save("img/{}2sobel_smooth.png".format(img[:-4]),dpi=(200,200))
    #extracted feature
    output_img_lap_sob=Image.fromarray(contrast_stretching(lap_sob))
    output_img_lap_sob.save("img/{}2lap_sob.png".format(img[:-4]),dpi=(200,200))
    #final imag
    output_img_final=Image.fromarray(contrast_stretching(img_final))
    output_img_final.save("img/{}2final.png".format(img[:-4]),dpi=(200,200))
    #final power low
    output_img_final_powerlow=Image.fromarray(contrast_stretching(img_final_powerlow))
    output_img_final_powerlow.save("img/{}2final_powerlow.png".format(img[:-4]),dpi=(200,200))

allworkinIMG(kid,1,0.8)
allworkinIMG(fruit,2,0.7)